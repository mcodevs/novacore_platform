"""Excel eksport — import YO'Q, faqat eksport (docs/04-flows/03-payroll-and-reports.md §6).

⚠️ `periods` / `payouts` yo'q (ADR-0015). Eksport **sana oralig'i** bo'yicha
ishlaydi (`frm` … `to`), pul kesimi esa qarz daftaridan olinadi:
`submissions.payable_amount − submissions.paid_amount`.
"""

from __future__ import annotations

import datetime as dt
import io
from dataclasses import dataclass
from decimal import Decimal

import sqlalchemy as sa
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import TASHKENT
from app.db.base import ZERO, as_utc, money, utcnow
from app.db.models import (
    Approval,
    ApprovalDecision,
    Employee,
    Submission,
    SubmissionLine,
    Vehicle,
)
from app.domain.payment import service as payment_service
from app.domain.stats import service as stats_service

HEADER_FONT = Font(bold=True)
MONEY_FMT = "#,##0"
#: To'lovlar tarixida ko'rsatiladigan eng ko'p yozuv
PAYMENTS_LIMIT = 1000


def _autosize(ws) -> None:  # noqa: ANN001
    for column in ws.columns:
        width = max((len(str(cell.value or "")) for cell in column), default=10)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 10), 45)


def _write_header(ws, headers: list[str]) -> None:  # noqa: ANN001
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _fmt_dt(value) -> str:  # noqa: ANN001
    if value is None:
        return ""
    return as_utc(value).astimezone(TASHKENT).strftime("%d.%m.%Y %H:%M")


def _date_slug(value: dt.datetime | None, *, fallback: str) -> str:
    """Fayl nomi uchun `YYYYMMDD` (Toshkent vaqti) yoki mazmunli chegara."""
    if value is None:
        return fallback
    return as_utc(value).astimezone(TASHKENT).strftime("%Y%m%d")


def _range_slug(frm: dt.datetime | None, to: dt.datetime | None) -> str:
    left = _date_slug(frm, fallback="boshidan")
    right = _date_slug(to, fallback="hozirgacha")
    return f"{left}_{right}"


def _save(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Ta'mirlar ----------------------------------------------------------------


async def export_submissions(
    session: AsyncSession,
    frm: dt.datetime | None = None,
    to: dt.datetime | None = None,
) -> tuple[str, bytes]:
    """`tamirlar_<from>_<to>.xlsx` — hisobotlar, to'liq ma'lumot + qarz holati."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ta'mirlar"
    # ⚠️ «So'ralgan ish haqi» va «Kamaytirildi» ustunlari ataylab YO'Q
    # (2026-08-05, ADR-0019 ruhi): hisobotning mavzusi — bajarilgan ish va
    # qarz, savdolashish emas. Kelishuv raqamlari kerak bo'lsa — «Ish
    # qatorlari» varag'ida qator kesimida turibdi.
    _write_header(
        ws,
        [
            "Raqam",
            "Holat",
            "Xodim",
            "Mashina",
            "Keldi",
            "Ketdi",
            "Downtime (soat)",
            "Tasdiqlangan ish haqi",
            "Qismlar",
            "Jami",
            "Qarz asosi",
            "To'langan",
            "Qolgan qarz",
            "Avtomatik tasdiq",
            "Yuborilgan",
        ],
    )

    stmt = (
        sa.select(Submission)
        .where(*stats_service.in_range(frm, to))
        .order_by(Submission.submitted_at)
    )
    for sub in (await session.execute(stmt)).scalars().all():
        author = await session.get(Employee, sub.author_id)
        vehicle = (
            await session.get(Vehicle, sub.subject_vehicle_id)
            if sub.subject_vehicle_id
            else None
        )
        downtime = sub.downtime_seconds
        approved = sub.labor_amount if sub.labor_amount is not None else None
        ws.append(
            [
                sub.number,
                sub.status.value,
                author.full_name if author else "",
                vehicle.plate_display if vehicle else "",
                _fmt_dt(sub.arrived_at),
                _fmt_dt(sub.left_at),
                round(downtime / 3600, 1) if downtime is not None else "",
                float(approved) if approved is not None else "",
                float(sub.parts_amount),
                float(sub.total_amount),
                float(sub.payable_amount),
                float(sub.paid_amount),
                float(sub.debt),
                "ha" if sub.auto_approved else "",
                _fmt_dt(sub.submitted_at),
            ]
        )

    #  Pul ustunlari: «Tasdiqlangan ish haqi» … «Qolgan qarz» (8–13).
    #  ⚠️ Ilgari bu oraliq bir ustunga surilgan edi (9–16): birinchi pul
    #  ustuni formatsiz qolib, matnli «Avtomatik tasdiq» ga format berilardi.
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=13):
        for cell in row:
            cell.number_format = MONEY_FMT
    _autosize(ws)

    # Ish qatorlari alohida varaqda.
    # ⚠️ «So'ralgan», «Kamaytirish sababi», «Rozilik» ustunlari ataylab YO'Q
    # (2026-08-05): varaq savdolashish jurnali emas, **bajarilgan ish** ro'yxati.
    # «Mashina» ustuni qo'shildi — mashina kesimida filtrlash uchun.
    ws2 = wb.create_sheet("Ish qatorlari")
    _write_header(
        ws2,
        [
            "Hisobot",
            "Mashina",
            "Xodim",
            "Tur",
            "Nomi",
            "Soni",
            "Tasdiqlangan",
            "O'z hisobidan",
        ],
    )
    line_stmt = (
        sa.select(SubmissionLine, Submission)
        .join(Submission, Submission.id == SubmissionLine.submission_id)
        .where(*stats_service.in_range(frm, to))
        .order_by(Submission.number)
    )
    for line, sub in (await session.execute(line_stmt)).all():
        author = await session.get(Employee, sub.author_id)
        line_vehicle = (
            await session.get(Vehicle, sub.subject_vehicle_id)
            if sub.subject_vehicle_id
            else None
        )
        ws2.append(
            [
                sub.number,
                line_vehicle.plate_display if line_vehicle else "",
                author.full_name if author else "",
                line.kind.value,
                line.name,
                float(line.qty),
                float(line.approved_amount) if line.approved_amount is not None else "",
                "ha" if line.self_funded else "",
            ]
        )
    #  Yagona pul ustuni — «Tasdiqlangan» (7). «Soni» (6) pul emas.
    for row in ws2.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = MONEY_FMT
    _autosize(ws2)

    return f"tamirlar_{_range_slug(frm, to)}.xlsx", _save(wb)


# --- Qarzlar ------------------------------------------------------------------


async def export_debts(
    session: AsyncSession,
    frm: dt.datetime | None = None,
    to: dt.datetime | None = None,
) -> tuple[str, bytes]:
    """`qarzlar_<sana>.xlsx` — buxgalteriyaga.

    Birinchi varaq — **hozirgi** holat (sana oralig'iga bog'liq emas: qarz
    yopilmaguncha ochiq turadi). Ikkinchi varaq — to'lovlar tarixi; oraliq
    berilgan bo'lsa, faqat shu oraliqdagi to'lovlar.
    """
    summary = await payment_service.debt_summary(session)

    wb = Workbook()
    ws = wb.active
    ws.title = "Qarzlar"
    ws.append(["Qarzdorlar — hozirgi holat"])
    ws["A1"].font = Font(bold=True, size=14)
    #  ⚠️ Manfiy «Sof qarz» — xato emas, holat: xodimda bizning pulimiz
    #  turibdi. Izohsiz buxgalter uni xato deb o'ylashi mumkin.
    ws.append(["Sof qarz = Qarz − Avans. Manfiy son — xodimda bizning pulimiz turibdi (P7)."])
    ws.append(["Xodim", "Ishlar soni", "Qarz", "Avans", "Sof qarz"])
    for cell in ws[3]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    #  ⚠️ Avansi bor, lekin qarzi yo'q xodim ham shu ro'yxatda (P7) — unda
    #  `count = 0`, `debt = 0`. Ilgari bunday qator «0 · 0» bo'lib turardi va
    #  ma'nosiz ko'rinardi; endi «Avans» ustuni uni tushuntiradi.
    for entry in summary.employees:
        net = entry.debt - entry.advance
        ws.append(
            [
                entry.full_name,
                entry.count,
                float(entry.debt),
                float(entry.advance),
                float(net),
            ]
        )
    ws.append([])
    ws.append(
        [
            "JAMI",
            "",
            float(summary.total),
            float(summary.advance_total),
            float(summary.total - summary.advance_total),
        ]
    )
    ws[f"A{ws.max_row}"].font = HEADER_FONT
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = MONEY_FMT
    _autosize(ws)

    # To'lovlar tarixi
    ws2 = wb.create_sheet("To'lovlar")
    _write_header(
        ws2,
        [
            "Sana",
            "Xodim",
            "Summa",
            "Izoh",
            "Hisobotlar",
            "Kiritdi",
            "Bekor qilingan",
            "Bekor sababi",
        ],
    )

    payments = await payment_service.payments_of(session, limit=PAYMENTS_LIMIT)
    if frm is not None:
        payments = [p for p in payments if as_utc(p.created_at) >= frm]
    if to is not None:
        payments = [p for p in payments if as_utc(p.created_at) <= to]

    # Allokatsiyalardagi hisobot raqamlari — bitta so'rovda
    numbers: dict[int, str] = {}
    ids = {alloc.submission_id for payment in payments for alloc in payment.allocations}
    if ids:
        rows = (
            await session.execute(
                sa.select(Submission.id, Submission.number).where(Submission.id.in_(ids))
            )
        ).all()
        numbers = {sub_id: number for sub_id, number in rows}

    total_paid = ZERO
    for payment in payments:
        if not payment.is_voided:
            total_paid = money(total_paid + payment.amount)
        actor = await session.get(Employee, payment.actor_id)
        ws2.append(
            [
                _fmt_dt(payment.created_at),
                payment.employee.full_name if payment.employee else "",
                float(payment.amount),
                payment.note or "",
                ", ".join(
                    numbers.get(alloc.submission_id, str(alloc.submission_id))
                    for alloc in payment.allocations
                ),
                actor.full_name if actor else "",
                "ha" if payment.is_voided else "",
                payment.void_reason or "",
            ]
        )
    ws2.append([])
    ws2.append(["JAMI (bekor qilinmagan)", "", float(total_paid)])
    ws2[f"A{ws2.max_row}"].font = HEADER_FONT
    for row in ws2.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = MONEY_FMT
    _autosize(ws2)

    today = _date_slug(utcnow(), fallback="")
    return f"qarzlar_{today}.xlsx", _save(wb)


# --- Kelishuv tejamkorligi ----------------------------------------------------


@dataclass
class _AuthorStat:
    """Xodim kesimi — `export_savings` ikkinchi varag'i."""

    full_name: str
    count: int = 0
    proposed: Decimal = ZERO
    approved: Decimal = ZERO
    debt: Decimal = ZERO
    paid: Decimal = ZERO
    disputes: int = 0

    @property
    def reduction_pct(self) -> Decimal:
        if self.proposed <= ZERO:
            return ZERO
        return money((self.proposed - self.approved) * 100 / self.proposed)


async def _author_breakdown(
    session: AsyncSession, frm: dt.datetime | None, to: dt.datetime | None
) -> list[_AuthorStat]:
    """Tasdiqlangan hisobotlarni muallif bo'yicha yig'adi (RPS past — Python'da)."""
    stmt = sa.select(Submission).where(
        *stats_service.in_range(frm, to),
        Submission.status.in_(stats_service.PAYABLE),
    )
    stats: dict[int, _AuthorStat] = {}
    for sub in (await session.execute(stmt)).scalars().all():
        stat = stats.get(sub.author_id)
        if stat is None:
            author = await session.get(Employee, sub.author_id)
            stat = _AuthorStat(
                full_name=author.full_name if author else str(sub.author_id)
            )
            stats[sub.author_id] = stat
        stat.count += 1
        stat.proposed = money(stat.proposed + sub.proposed_labor_amount)
        stat.approved = money(stat.approved + (sub.labor_amount or ZERO))
        stat.paid = money(stat.paid + sub.paid_amount)
        stat.debt = money(stat.debt + (sub.payable_amount - sub.paid_amount))

    disputes_stmt = (
        sa.select(Submission.author_id, sa.func.count(Approval.id))
        .join(Approval, Approval.submission_id == Submission.id)
        .where(
            *stats_service.in_range(frm, to),
            Approval.decision == ApprovalDecision.price_disputed,
        )
        .group_by(Submission.author_id)
    )
    for author_id, count in (await session.execute(disputes_stmt)).all():
        if author_id in stats:
            stats[author_id].disputes = int(count)

    return sorted(stats.values(), key=lambda item: item.approved, reverse=True)


async def export_savings(
    session: AsyncSession,
    frm: dt.datetime | None = None,
    to: dt.datetime | None = None,
) -> tuple[str, bytes]:
    """⭐ `kelishuv_<from>_<to>.xlsx` — narx kelishuvi tejamkorligi."""
    summary = await stats_service.range_summary(session, frm=frm, to=to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kelishuv"
    ws.append([f"Narx kelishuvi — {summary.title}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    rows: list[tuple[str, object, bool]] = [
        ("Ustalar so'radi", float(summary.proposed_total), True),
        ("Tasdiqlandi", float(summary.approved_total), True),
        ("TEJALDI", float(summary.saved), True),
        ("Tejamkorlik %", float(summary.saved_pct), False),
        ("Qismlar", float(summary.parts_total), True),
        ("Jami hisobotlar", summary.total_submissions, False),
        ("Tasdiqlangan hisobotlar", summary.approved_count, False),
        ("Avtomatik tasdiqlangan (admin)", summary.auto_approved_count, False),
        ("Avtomatik tasdiqlangan summa", float(summary.auto_approved_total), True),
        ("To'langan", float(summary.paid_total), True),
        ("Qolgan qarz", float(summary.debt_total), True),
    ]
    for label, value, is_money in rows:
        ws.append([label, value])
        if is_money:
            ws.cell(row=ws.max_row, column=2).number_format = MONEY_FMT
    _autosize(ws)

    ws2 = wb.create_sheet("Xodimlar kesimi")
    _write_header(
        ws2,
        [
            "Xodim",
            "Ishlar",
            "So'radi",
            "Tasdiqlandi",
            "Kamaytirish %",
            "To'langan",
            "Qolgan qarz",
            "Nizolar",
        ],
    )
    for stat in await _author_breakdown(session, frm, to):
        ws2.append(
            [
                stat.full_name,
                stat.count,
                float(stat.proposed),
                float(stat.approved),
                float(stat.reduction_pct),
                float(stat.paid),
                float(stat.debt),
                stat.disputes,
            ]
        )
    for row in ws2.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = MONEY_FMT
    for row in ws2.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = MONEY_FMT
    _autosize(ws2)

    return f"kelishuv_{_range_slug(frm, to)}.xlsx", _save(wb)


EXPORTS = {
    "submissions": export_submissions,
    "debts": export_debts,
    "savings": export_savings,
}


async def build(
    session: AsyncSession,
    kind: str,
    frm: dt.datetime | None = None,
    to: dt.datetime | None = None,
) -> tuple[str, bytes]:
    if kind not in EXPORTS:
        raise ValueError(f"unknown export: {kind}")
    return await EXPORTS[kind](session, frm=frm, to=to)


__all__ = ["build", "export_debts", "export_savings", "export_submissions"]
