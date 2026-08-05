"""Excel eksport — ustunlar tarkibi.

⚠️ Bu yerda **sarlavha qatori qotirib qo'yilgan**. Sabab: eksport hech qayerda
ko'rinmaydi — xato ustun yoki surilib ketgan format faqat buxgalter faylni
ochganda bilinadi. Ustun qo'shish/olib tashlash — ataylab qilinadigan ish,
shuning uchun u testni ham o'zgartirishni talab qilsin.
"""

from __future__ import annotations

import io
from decimal import Decimal

import openpyxl

from app.db.models import Employee, SubmissionStatus
from app.domain.approval import service as approval_service
from app.domain.export import service as export_service
from app.domain.payment import service as payment_service
from app.domain.pricing import service as pricing_service
from app.domain.submission import service as submission_service
from tests.conftest import create_ready_submission, make_employee, make_vehicle

#: «So'ralgan ish haqi» va «Kamaytirildi» ataylab YO'Q (2026-08-05, ADR-0019 ruhi)
REPAIRS_HEADER = [  # noqa: RUF012
    "Raqam",
    "Holat",
    "Xodim",
    "Mashina",
    "Keldi",
    "Ketdi",
    "Downtime (soat)",
    "Tasdiqlangan ish haqi",
    "Qismlar",
    "Jami",
    "Qarz asosi",
    "To'langan",
    "Qolgan qarz",
    "Avtomatik tasdiq",
    "Yuborilgan",
]

#: «So'ralgan», «Kamaytirish sababi», «Rozilik» ataylab YO'Q; «Mashina» qo'shilgan
LINES_HEADER = [  # noqa: RUF012
    "Hisobot",
    "Mashina",
    "Xodim",
    "Tur",
    "Nomi",
    "Soni",
    "Tasdiqlangan",
    "O'z hisobidan",
]


def _sheet(payload: bytes, title: str):  # noqa: ANN202
    return openpyxl.load_workbook(io.BytesIO(payload))[title]


async def _approved(session):  # noqa: ANN001, ANN202
    mechanic = await make_employee(session, role_code="mechanic")
    admin = await make_employee(session, role_code="admin", name="Admin A.")
    vehicle = await make_vehicle(session)
    submission = await create_ready_submission(
        session, mechanic, vehicle, works=[("Kolodka almashtirish", Decimal("250000"))]
    )
    await submission_service.submit(session, submission, mechanic)
    line = submission.lines[0]
    await pricing_service.propose_price(
        session, submission, admin, changes=[(line.id, Decimal("200000"))], comment="tarixga ko'ra"
    )
    await pricing_service.accept_price(session, submission, mechanic)
    assert submission.status == SubmissionStatus.APPROVED
    return submission


async def test_repairs_sheet_has_no_negotiation_columns(session):
    """⭐ Egasining talabi: hisobotda «So'ralgan» va «Kamaytirildi» ustunlari yo'q."""
    await _approved(session)
    _name, payload = await export_service.export_submissions(session)
    ws = _sheet(payload, "Ta'mirlar")

    header = [cell.value for cell in ws[1]]
    assert header == REPAIRS_HEADER
    assert "So'ralgan ish haqi" not in header
    assert "Kamaytirildi" not in header


async def test_repairs_sheet_values_match_header(session):
    """Ustun olib tashlanganda qatorlar surilib ketmasin — qiymat sarlavhaga mos."""
    submission = await _approved(session)
    _name, payload = await export_service.export_submissions(session)
    ws = _sheet(payload, "Ta'mirlar")

    row = dict(zip(REPAIRS_HEADER, [cell.value for cell in ws[2]], strict=True))
    assert row["Raqam"] == submission.number
    assert row["Tasdiqlangan ish haqi"] == 200000  # kamaytirilgan summa
    assert row["Jami"] == 200000
    assert row["Qarz asosi"] == float(submission.payable_amount)
    assert row["Qolgan qarz"] == float(submission.debt)


async def test_money_format_covers_exactly_the_money_columns(session):
    """Format oralig'i sarlavha bilan birga surilishi kerak (ilgari 1 ga siljigan edi)."""
    await _approved(session)
    _name, payload = await export_service.export_submissions(session)
    ws = _sheet(payload, "Ta'mirlar")

    money_columns = {
        "Tasdiqlangan ish haqi",
        "Qismlar",
        "Jami",
        "Qarz asosi",
        "To'langan",
        "Qolgan qarz",
    }
    for index, title in enumerate(REPAIRS_HEADER, start=1):
        formatted = ws.cell(row=2, column=index).number_format == "#,##0"
        assert formatted is (title in money_columns), title


async def test_lines_sheet_is_a_work_list_not_a_haggling_log(session):
    """⭐ Egasining talabi: qatorlar varag'i — bajarilgan ish ro'yxati.

    «So'ralgan», «Kamaytirish sababi», «Rozilik» olib tashlandi; o'rniga
    mashina raqami qo'shildi (mashina kesimida filtrlash uchun).
    """
    submission = await _approved(session)
    _name, payload = await export_service.export_submissions(session)
    ws = _sheet(payload, "Ish qatorlari")

    header = [cell.value for cell in ws[1]]
    assert header == LINES_HEADER
    for gone in ("So'ralgan", "Kamaytirish sababi", "Rozilik"):
        assert gone not in header

    row = dict(zip(LINES_HEADER, [cell.value for cell in ws[2]], strict=True))
    assert row["Hisobot"] == submission.number
    assert row["Mashina"] == submission.vehicle.plate_display
    assert row["Tasdiqlangan"] == 200000
    assert ws.cell(row=2, column=header.index("Tasdiqlangan") + 1).number_format == "#,##0"
    # «Soni» — pul emas, format berilmasin
    assert ws.cell(row=2, column=header.index("Soni") + 1).number_format != "#,##0"


async def test_auto_approved_report_exports(session):
    """R1a — avtomatik tasdiqlangan hisobot ham eksportga tushadi."""
    admin = await make_employee(session, role_code="admin", name="Admin B.")
    vehicle = await make_vehicle(session, plate="01B456CD")
    submission = await create_ready_submission(
        session, admin, vehicle, works=[("Diagnostika", Decimal("120000"))]
    )
    await submission_service.submit(session, submission, admin)

    _name, payload = await export_service.export_submissions(session)
    ws = _sheet(payload, "Ta'mirlar")
    row = dict(zip(REPAIRS_HEADER, [cell.value for cell in ws[2]], strict=True))
    assert row["Avtomatik tasdiq"] == "ha"
    assert row["Tasdiqlangan ish haqi"] == 120000


async def test_export_filename_uses_range(session):
    _name, _payload = await export_service.export_submissions(session)
    assert _name.startswith("tamirlar_") and _name.endswith(".xlsx")


# --- Qarzlar eksporti ---------------------------------------------------------


async def _overpaid(session):  # noqa: ANN001, ANN202
    """Qarzidan ko'p to'langan xodim — ortig'i avans bo'lib qoladi (P7)."""
    submission = await _approved(session)
    await session.refresh(submission)
    mechanic_id = submission.author_id
    admin = await make_employee(session, role_code="admin", name="Buxgalter B.")
    await payment_service.create_payment(
        session,
        employee_id=mechanic_id,
        actor_id=admin.id,
        amount=submission.payable_amount + Decimal("50000"),
    )
    return mechanic_id


async def test_debts_export_has_advance_sheet(session):
    """⭐ Egasining talabi: xodimdagi avans ham shu hujjatda ko'rinsin."""
    await _overpaid(session)
    _name, payload = await export_service.export_debts(session)
    wb = openpyxl.load_workbook(io.BytesIO(payload))

    assert wb.sheetnames == ["Qarzlar", "Avans", "To'lovlar"]
    ws = wb["Avans"]
    assert [cell.value for cell in ws[3]] == ["Xodim", "Avans"]

    rows = [(r[0], r[1]) for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]
    names = {name for name, _ in rows if name != "JAMI"}
    assert names  # avansi bor xodim bor
    total = next(value for name, value in rows if name == "JAMI")
    assert total == 50000
    assert ws.cell(row=4, column=2).number_format == "#,##0"


async def test_debtors_sheet_skips_advance_only_employee(session):
    """Avansi bor, qarzi yo'q xodim qarzdorlar jadvalida «0 · 0» bo'lib turmasin."""
    await _overpaid(session)
    _name, payload = await export_service.export_debts(session)
    ws = openpyxl.load_workbook(io.BytesIO(payload))["Qarzlar"]

    body = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0] and r[0] != "JAMI"]
    assert body == []  # yagona xodimning qarzi yopilgan → qarzdor yo'q

    total = next(r[2] for r in ws.iter_rows(min_row=4, values_only=True) if r[0] == "JAMI")
    assert total == 0


async def test_advance_is_spent_and_shows_as_reduced_debt(session):
    """P7 eksportda: avans yangi qarzga ishlatiladi → «Avans» varag'idan chiqadi,
    «Qarzlar» da esa qarz ayrilgan holda turadi."""
    mechanic_id = await _overpaid(session)
    mechanic = await session.get(Employee, mechanic_id)
    vehicle = await make_vehicle(session, plate="01C789DE")
    second = await create_ready_submission(
        session, mechanic, vehicle, works=[("Yangi ish", Decimal("300000"))]
    )
    await submission_service.submit(session, second, mechanic)
    admin = await make_employee(session, role_code="admin", name="Admin C.")
    await approval_service.approve(session, second, admin)
    await session.refresh(second)
    # P7 — tasdiqlanganda avans avtomatik ishlatiladi
    assert second.paid_amount == Decimal("50000.00")

    _name, payload = await export_service.export_debts(session)
    wb = openpyxl.load_workbook(io.BytesIO(payload))

    debtors = [r for r in wb["Qarzlar"].iter_rows(min_row=4, values_only=True) if r[0] == mechanic.full_name]
    assert debtors and debtors[0][2] == 250000  # 300 000 − 50 000 avans

    advances = [r for r in wb["Avans"].iter_rows(min_row=4, values_only=True) if r[0] == mechanic.full_name]
    assert advances == []  # avans sarflandi
